import copy
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Helper constants
DAYS = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
DEFAULT_END_TIME = "20h00"



filter_settings = {
    "free_days": False,
    "end_time": DEFAULT_END_TIME,
    "professor": ""
}


# As horas têm de ter todas 5 digitos
aulas = {
    "AED": [
        [(1, "Quinta", "17:30", "19:30", ["PMPR", "PNFRCD"], "T"),
         (1, "Terça", "08:30", "10:30", ["DEI_2_AED_2"], "TP")],
        [(2, "Quinta", "17:30", "19:30", ["PMPR", "PNFRCD"], "T"),
         (2, "Segunda", "16:00", "19:00", ["FMMR"], "TP")],
        [(3, "Quinta", "17:30", "19:30", ["PMPR", "PNFRCD"], "T"),
         (3, "Quarta", "11:00", "13:00", ["DEI_2_AED_2"], "TP")],
        [(4, "Quinta", "17:30", "19:30", ["PMPR", "PNFRCD"], "T"),
         (4, "Quarta", "10:30", "12:30", ["BJCL"], "TP")],
        [(5, "Quinta", "17:30", "19:30", ["PMPR", "PNFRCD"], "T"),
         (5, "Terça", "08:30", "10:30", ["DEI_2_AED_1"], "TP")],
        [(6, "Quinta", "17:30", "19:30", ["PMPR", "PNFRCD"], "T"),
         (6, "Sexta", "17:00", "19:00", ["VAFS"], "TP")],
        [(7, "Quinta", "17:30", "19:30", ["PMPR", "PNFRCD"], "T"),
         (7, "Sexta", "17:00", "19:00", ["IXSDS"], "TP")],
        [(8, "Quinta", "17:30", "19:30", ["PMPR", "PNFRCD"], "T"),
         (8, "Quinta", "14:00", "16:00", ["APR"], "TP")],
        [(9, "Quarta", "10:30", "12:30", ["PMPR", "PNFRCD"], "T"),
         (9, "Sexta", "14:00", "16:00", ["IXSDS"], "TP")],
        [(10, "Quarta", "10:30", "12:30", ["PMPR", "PNFRCD"], "T"),
         (10, "Quarta", "08h30", "10h30", ["BJCL"], "TP")],
        [(11, "Quarta", "10:30", "12:30", ["PMPR", "PNFRCD"], "T"),
         (11, "Segunda", "14h00", "16h00", ["FMMR"], "TP")],
        [(12, "Quarta", "10:30", "12:30", ["PMPR", "PNFRCD"], "T"),
         (12, "Terça", "10h30", "12h30", ["DEI_2_AED_1"], "TP")],
        [(13, "Quarta", "10:30", "12:30", ["PMPR", "PNFRCD"], "T"),
         (13, "Quarta", "08h30", "10h30", ["BJCL"], "TP")],
        [(14, "Quarta", "10:30", "12:30", ["PMPR", "PNFRCD"], "T"),
         (14, "Quinta", "16h00", "18h00", ["APR"], "TP")],
        [(15, "Quarta", "10:30", "12:30", ["PMPR", "PNFRCD"], "T"),
         (15, "Sexta", "14h00", "16h00", ["VAFS"], "TP")],
        [(16, "Quarta", "10:30", "12:30", ["PMPR", "PNFRCD"], "T"),
         (16, "Sexta", "14h00", "16h00", ["VAFS"], "TP")],
    ],
    "BD": [
        [(1,"Segunda", "13:00", "14:00", ["MCPF", "CTL"], "T"),
         (1,"Sexta", "16:00", "17:00", ["MCPF", "CTL"], "T"),
         (1,"Quarta", "10:30", "12:30", ["ANE"], "TP"),],
        
        [(2,"Segunda", "13:00", "14:00", ["MCPF", "CTL"], "T"),
         (2,"Sexta", "16:00", "17:00", ["MCPF", "CTL"], "T"),
         (2,"Sexta", "17h00", "19h00", ["JPMD"], "TP"),],
        
        [(3,"Segunda", "13:00", "14:00", ["MCPF", "CTL"], "T"),
         (3,"Sexta", "16:00", "17:00", ["MCPF", "CTL"], "T"),
         (3,"Quinta", "14h00", "16h00", ["DFG"], "TP"),],
        
        [(4,"Segunda", "13:00", "14:00", ["MCPF", "CTL"], "T"),
         (4,"Sexta", "16:00", "17:00", ["MCPF", "CTL"], "T"),
         (4,"Segunda", "16h00", "18h00", ["LGBC"], "TP"),],
        
        [(5,"Segunda", "13:00", "14:00", ["MCPF", "CTL"], "T"),
         (5,"Sexta", "16:00", "17:00", ["MCPF", "CTL"], "T"),
         (5,"Sexta", "17h00", "19h00", ["MFD"], "TP"),],
        
        [(6,"Segunda", "13:00", "14:00", ["MCPF", "CTL"], "T"),
         (6,"Sexta", "16:00", "17:00", ["MCPF", "CTL"], "T"),
         (6,"Segunda", "16h00", "18h00", ["MCPF"], "TP"),],
        
        [(7,"Segunda", "13:00", "14:00", ["MCPF", "CTL"], "T"),
         (7,"Sexta", "16:00", "17:00", ["MCPF", "CTL"], "T"),
         (7,"Terça", "08h30", "10h30", ["DEI_PAUX_SINF"], "TP"),],
        
        [(8,"Segunda", "13:00", "14:00", ["MCPF", "CTL"], "T"),
         (8,"Sexta", "16:00", "17:00", ["MCPF", "CTL"], "T"),
         (8,"Terça", "08h30", "10h30", ["ASP"], "TP"),],
        
        [(9,"Segunda", "18h00", "19h00", ["MCPF", "CTL"], "T"),
         (9,"Sexta", "18:00", "19:00", ["MCPF", "CTL"], "T"),
         (9,"Quinta", "16h00", "18h00", ["DFG"], "TP"),],
        
        [(10,"Segunda", "18h00", "19h00", ["MCPF", "CTL"], "T"),
         (10,"Sexta", "18:00", "19:00", ["MCPF", "CTL"], "T"),
         (10,"Terça", "10h30", "12h30", ["ASP"], "TP"),],
        
        [(11,"Segunda", "18h00", "19h00", ["MCPF", "CTL"], "T"),
         (11,"Sexta", "18:00", "19:00", ["MCPF", "CTL"], "T"),
         (11,"Quinta", "16h00", "18h00", ["AHM"], "TP"),],
        
        [(12,"Segunda", "18h00", "19h00", ["MCPF", "CTL"], "T"),
         (12,"Sexta", "18:00", "19:00", ["MCPF", "CTL"], "T"),
         (12,"Sexta", "14h00", "16h00", ["MFD"], "TP"),],
        
        [(13,"Segunda", "18h00", "19h00", ["MCPF", "CTL"], "T"),
         (13,"Sexta", "18:00", "19:00", ["MCPF", "CTL"], "T"),
         (13,"Segunda", "14h00", "16h00", ["MCPF"], "TP"),],
        
        [(14,"Segunda", "18h00", "19h00", ["MCPF", "CTL"], "T"),
         (14,"Sexta", "18:00", "19:00", ["MCPF", "CTL"], "T"),
         (14,"Segunda", "14h00", "16h00", ["LGBC"], "TP"),],
        
        [(15,"Segunda", "18h00", "19h00", ["MCPF", "CTL"], "T"),
         (15,"Sexta", "18:00", "19:00", ["MCPF", "CTL"], "T"),
         (15,"Quarta", "08h30", "10h30", ["ANE"], "TP"),],
        
        [(16,"Segunda", "18h00", "19h00", ["MCPF", "CTL"], "T"),
         (16,"Sexta", "18:00", "19:00", ["MCPF", "CTL"], "T"),
         (16,"Quarta", "08h00", "10h00", ["CTL"], "TP"),],    
    ],
    
    "FII": [
        [(1, "Quinta", "16:00", "17:30", ["JCREO"], "T"),
         (1, "Quinta", "14:00", "15:30", ["MFGM"], "TP")],
        [(2, "Quinta", "16:00", "17:30", ["JCREO"], "T"),
         (2, "Quinta", "14:30", "16:00", ["JB"], "TP")],
        [(3, "Quinta", "16:00", "17:30", ["JCREO"], "T"),
         (3, "Segunda", "16:00", "17:30", ["JCREO"], "TP")],
        [(4, "Quinta", "16:00", "17:30", ["JCREO"], "T"),
         (4, "Sexta", "17:00", "18:30", ["AGCG"], "TP")],
        [(5, "Quinta", "16:00", "17:30", ["JCREO"], "T"),
         (5, "Segunda", "16:00", "17:30", ["PPA"], "TP")],
        [(6, "Quinta", "16:00", "17:30", ["JCREO"], "T"),
         (6, "Quinta", "14:30", "16:00", ["DPU"], "TP")],
        [(7, "Quinta", "16:00", "17:30", ["JCREO"], "T"),
         (7, "Quinta", "14:00", "15:30", ["DEI_1_FP_1"], "TP")],
        [(8, "Quinta", "16:00", "17:30", ["JCREO"], "T"),
         (8, "Sexta", "17:00", "18:30", ["JCREO"], "TP")],
        [(9, "Quinta", "14:30", "16:00", ["JCREO"], "T"),
         (9, "Quarta", "09:00", "10:30", ["DPU"], "TP")],
        [(10, "Quinta", "14:30", "16:00", ["JCREO"], "T"),
         (10, "Sexta", "14:00", "15:30", ["JCREO"], "TP")],
        [(11, "Quinta", "14:30", "16:00", ["JCREO"], "T"),
         (11, "Terça", "10:00", "11:30", ["DEI_1_FP_2"], "TP")],
        [(12, "Quinta", "14:30", "16:00", ["JCREO"], "T"),
         (12, "Quinta", "16:00", "17:30", ["JB"], "TP")],
        [(13, "Quinta", "14:30", "16:00", ["JCREO"], "T"),
         (13, "Sexta", "14:30", "16:00", ["AGCG"], "TP")],
        [(14, "Quinta", "14:30", "16:00", ["JCREO"], "T"),
         (14, "Quinta", "11:00", "12:30", ["DPU"], "TP")],
        [(15, "Quinta", "14:30", "16:00", ["JCREO"], "T"),
         (15, "Segunda", "14:30", "16:00", ["PPA"], "TP")],
        [(16, "Quinta", "14:30", "16:00", ["JCREO"], "T"),
         (16, "Segunda", "14:30", "16:00", ["PPA"], "TP")],
    ],
    
    "LDTS" : [
        [(1, "Segunda", "14h00", "16h00", ["rma"], "T"),
         (1, "Segunda", "16h00", "18h00", ["AOR"], "PL")],
        [(2, "Segunda", "14h00", "16h00", ["rma"], "T"),
         (2, "Terça", "08h30", "10h30", ["rma"], "PL")],
        [(3, "Segunda", "14h00", "16h00", ["rma"], "T"),
         (3, "Sexta", "17h00", "19h00", ["JAC"], "PL")],
        [(4, "Segunda", "14h00", "16h00", ["rma"], "T"),
         (4, "Quinta", "14h00", "16h00", ["JCMC"], "PL")],
        [(5, "Segunda", "14h00", "16h00", ["rma"], "T"),
         (5, "Quarta", "10h30", "12h30", ["LFFG"], "PL")],
        [(6, "Segunda", "14h00", "16h00", ["rma"], "T"),
         (6, "Terça", "08h30", "10h30", ["DEI_2_LDSO_1 - Sofia"], "PL")],
        [(7, "Segunda", "14h00", "16h00", ["rma"], "T"),
         (7, "Quarta", "10h30", "12h30", ["DABF"], "PL")],
        [(8, "Segunda", "14h00", "16h00", ["rma"], "T"),
         (8, "Quarta", "10h30", "12h30", ["DABF"], "PL")],
        [(9, "Segunda", "16h00", "18h00", ["rma"], "T"),
         (9, "Segunda", "14h00", "16h00", ["AOR"], "PL")],
        [(10, "Segunda", "16h00", "18h00", ["rma"], "T"),
         (10, "Quinta", "16h00", "18h00", ["AOR"], "PL")],
        [(11, "Segunda", "16h00", "18h00", ["rma"], "T"),
         (11, "Quarta", "08h30", "10h30", ["LFFG"], "PL")],
        [(12, "Segunda", "16h00", "18h00", ["rma"], "T"),
         (12, "Quarta", "08h30", "10h30", ["DABF"], "PL")],
        [(13, "Segunda", "16h00", "18h00", ["rma"], "T"),
         (13, "Quinta", "16h00", "18h00", ["JCMC"], "PL")],
        [(14, "Segunda", "16h00", "18h00", ["rma"], "T"),
         (14, "Sexta", "14h00", "16h00", ["JAC"], "PL")],
        [(15, "Segunda", "16h00", "18h00", ["rma"], "T"),
         (15, "Terça", "10h30", "12h30", ["DEI_2_LDSO_1 - Sofia"], "PL")],
        [(16, "Segunda", "16h00", "18h00", ["rma"], "T"),
         (16, "Terça", "10h30", "12h30", ["DEI_2_LDSO_1 - Sofia"], "PL")],
    ],
    
    "SO" : [
        [(1, "Sexta", "14h00", "16h00", ["LMBL", "CMFB-M"], "T"),
         (1, "Sexta", "17h00", "19h00", ["JHSO"], "TP")],
        [(2, "Sexta", "14h00", "16h00", ["LMBL", "CMFB-M"], "T"),
         (2, "Quarta", "10h30", "12h30", ["TLS"], "TP")],
        [(3, "Sexta", "14h00", "16h00", ["LMBL", "CMFB-M"], "T"),
         (3, "Terça", "08h30", "10h30", ["HMSO"], "TP")],
        [(4, "Sexta", "14h00", "16h00", ["LMBL", "CMFB-M"], "T"),
         (4, "Terça", "08h30", "10h30", ["LMBL"], "TP")],
        [(5, "Sexta", "14h00", "16h00", ["LMBL", "CMFB-M"], "T"),
         (5, "Quinta", "14h00", "16h00", ["LFOP"], "TP")],
        [(6, "Sexta", "14h00", "16h00", ["LMBL", "CMFB-M"], "T"),
         (6, "Quarta", "10h30", "12h30", ["MMC"], "TP")],
        [(7, "Sexta", "14h00", "16h00", ["LMBL", "CMFB-M"], "T"),
         (7, "Segunda", "16h00", "18h00", ["PRCS"], "TP")],
        [(8, "Sexta", "14h00", "16h00", ["LMBL", "CMFB-M"], "T"),
         (8, "Segunda", "16h00", "18h00", ["PRCS"], "TP")],
        [(9, "Sexta", "16h00", "18h00", ["LMBL", "CMFB-M"], "T"),
         (9, "Terça", "10h30", "12h30", ["HMSO"], "TP")],
        [(10, "Sexta", "16h00", "18h00", ["LMBL", "CMFB-M"], "T"),
         (10, "Segunda", "14h00", "16h00", ["FAFAM"], "TP")],
        [(11, "Sexta", "16h00", "18h00", ["LMBL", "CMFB-M"], "T"),
         (11, "Sexta", "14h00", "16h00", ["JHSO"], "TP")],
        [(12, "Sexta", "16h00", "18h00", ["LMBL", "CMFB-M"], "T"),
         (12, "Segunda", "14h00", "16h00", ["PRCS"], "TP")],
        [(13, "Sexta", "16h00", "18h00", ["LMBL", "CMFB-M"], "T"),
         (13, "Terça", "10h30", "12h30", ["LMBL"], "TP")],
        [(14, "Sexta", "16h00", "18h00", ["LMBL", "CMFB-M"], "T"),
         (14, "Quarta", "08h30", "10h30", ["MMC"], "TP")],
        [(15, "Sexta", "16h00", "18h00", ["LMBL", "CMFB-M"], "T"),
         (15, "Quinta", "16h00", "18h00", ["LFOP"], "TP")],
        [(16, "Sexta", "16h00", "18h00", ["LMBL", "CMFB-M"], "T"),
         (16, "Quinta", "16h00", "18h00", ["LFOP"], "TP")],
    ]
    
    
}

#         Seg Ter Qua Qui Sex
horario = [[], [], [], [], []]
def no_time_conflict(start1, end1, start2, end2):
    """
    Check if two time intervals conflict.
    """
    start1_hour, start1_minute = map(int, start1.replace('h', ':').split(':'))
    end1_hour, end1_minute = map(int, end1.replace('h', ':').split(':'))
    start2_hour, start2_minute = map(int, start2.replace('h', ':').split(':'))
    end2_hour, end2_minute = map(int, end2.replace('h', ':').split(':'))

    # Check no overlap conditions
    if (end1_hour < start2_hour) or (end1_hour == start2_hour and end1_minute <= start2_minute):
        return True
    if (end2_hour < start1_hour) or (end2_hour == start1_hour and end2_minute <= start1_minute):
        return True
    return False

def generate_schedules(courses, remaining_courses, current_schedule):
    """
    Recursively generate all possible schedules.
    """
    if not remaining_courses:
        return [copy.deepcopy(current_schedule)]

    schedules_list = []
    current_course = courses[remaining_courses[0]]
    
    for group in current_course:
        fits = True
        temp_schedule = copy.deepcopy(current_schedule)
        
        for session in group:
            day_index = DAYS.index(session[1])
            
            # Check for conflicts with already scheduled sessions
            for existing_session in temp_schedule[day_index]:
                if not no_time_conflict(session[2], session[3], existing_session[2], existing_session[3]):
                    fits = False
                    break
            
            if not fits:
                break
            
            # Add session to the temporary schedule
            temp_schedule[day_index].append((remaining_courses[0] + str(session[0]),) + session[1:])
        
        # Recursively continue with the next course
        if fits:
            schedules_list.extend(generate_schedules(courses, remaining_courses[1:], temp_schedule))

    return schedules_list

def output_schedules_to_excel(schedules, filename="schedules_all.xlsx"):
    """
    Outputs the given list of schedules to an Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Schedules"
    
    colors = {
        "AED": "FFC000",  # Orange
        "BD": "92D050",   # Green
        "FII": "00B0F0",  # Blue
        "LDTS": "FF00FF", # Pink
        "SO": "7030A0"    # Purple
    }

    time_slots = [
        "08:00-08:30", "08:30-09:00", "09:00-09:30", "09:30-10:00",
        "10:00-10:30", "10:30-11:00", "11:00-11:30", "11:30-12:00",
        "12:00-12:30", "12:30-13:00", "13:00-13:30", "13:30-14:00",
        "14:00-14:30", "14:30-15:00", "15:00-15:30", "15:30-16:00",
        "16:00-16:30", "16:30-17:00", "17:00-17:30", "17:30-18:00",
        "18:00-18:30", "18:30-19:00", "19:00-19:30", "19:30-20:00"
    ]
    
    def time_to_row(time):
        """
        Convert a time string (HH:MM or HHhMM) to a corresponding row in Excel.
        """
        hour, minute = map(int, time.replace('h', ':').split(':'))
        return (hour - 8) * 2 + (1 if minute >= 30 else 0) + 2

    def apply_cell_style(cell, fill_color="FFFFFF"):
        """
        Apply style to a given Excel cell.
        """
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    ws.column_dimensions['A'].width = 15
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

    for idx, schedule in enumerate(schedules):
        start_row = idx * (len(time_slots) + 3) + 1
        
        ws.cell(row=start_row, column=1, value=f"Schedule {idx + 1}")
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        apply_cell_style(ws.cell(row=start_row, column=1), "DDDDDD")
        
        ws.cell(row=start_row + 1, column=1, value="Hora")
        for col, day in enumerate(DAYS, start=2):
            ws.cell(row=start_row + 1, column=col, value=day)
            apply_cell_style(ws.cell(row=start_row + 1, column=col), "DDDDDD")
        
        for row, time_slot in enumerate(time_slots, start=start_row + 2):
            ws.cell(row=row, column=1, value=time_slot)
            apply_cell_style(ws.cell(row=row, column=1), "DDDDDD")
        
        for day_idx, day_schedule in enumerate(schedule):
            for class_info in day_schedule:
                full_course = class_info[0]
                course = ''.join([c for c in full_course if not c.isdigit()])
                start_time = class_info[2]
                end_time = class_info[3]
                class_type = class_info[5]
                room = ", ".join(class_info[4])
                
                start_row_offset = time_to_row(start_time)
                end_row_offset = time_to_row(end_time)
                
                for row_offset in range(start_row_offset, end_row_offset):
                    cell = ws.cell(row=start_row + row_offset, column=day_idx + 2)
                    cell.value = f"{full_course} ({class_type})\n{room}"
                    apply_cell_style(cell, colors.get(course, "FFFFFF"))
    
    wb.save(filename)
    print(f"Schedules have been saved to {filename}")

def filter_schedules_by_free_days(schedules):
    """
    Filters schedules to include only those that have at least one completely free day.
    """
    return [schedule for schedule in schedules if any(len(day) == 0 for day in schedule)]

def filter_schedules_by_end_time(schedules, end_time):
    """
    Filters schedules to include only those that end before a specified time.
    """
    def parse_time(time_str):
        """
        Parses a time string into hours and minutes.
        """
        if ':' in time_str:
            return map(int, time_str.split(':'))
        elif 'h' in time_str:
            return map(int, time_str.replace('h', ':').split(':'))
        else:
            raise ValueError(f"Unexpected time format: {time_str}. Expected format: 'HH:MM' or 'HHhMM'")

    try:
        end_hour, end_minute = parse_time(end_time)
    except (TypeError, ValueError) as e:
        print(f"Error parsing end time: {e}")
        return schedules
    
    def ends_before_time(schedule):
        """
        Checks if a schedule ends before a specified time.
        """
        for day in schedule:
            if day:
                last_class = max(day, key=lambda x: x[3])
                class_end_hour, class_end_minute = parse_time(last_class[3])
                
                if class_end_hour > end_hour or (class_end_hour == end_hour and class_end_minute > end_minute):
                    return False
        return True
    
    return list(filter(ends_before_time, schedules))

def filter_schedules_by_professor(schedules, professor):
    """
    Filters schedules to include only those that have at least one class with the specified professor.
    """
    filtered_schedules = []
    
    for schedule in schedules:
        contains_professor = False
        for day in schedule:
            for class_info in day:
                # class_info[4] is the list of professors for this class
                if professor in class_info[4]:
                    contains_professor = True
                    break
            if contains_professor:
                break
        
        if contains_professor:
            filtered_schedules.append(schedule)
    
    return filtered_schedules

def apply_filters(schedules, settings):
    """
    Applies all active filters to the schedules based on the current settings.
    """
    filtered_schedules = schedules

    if settings["free_days"]:
        filtered_schedules = filter_schedules_by_free_days(filtered_schedules)

    if settings["end_time"] != DEFAULT_END_TIME:
        filtered_schedules = filter_schedules_by_end_time(filtered_schedules, settings["end_time"])

    if settings["professor"]:
        filtered_schedules = filter_schedules_by_professor(filtered_schedules, settings["professor"])

    return filtered_schedules


def count_subject_days(schedule):
    """
    Counts the number of different days a subject appears in a given schedule.
    """
    subject_days = {}
    
    for day_index, day_schedule in enumerate(schedule):
        for class_info in day_schedule:
            subject = class_info[0]
            if subject not in subject_days:
                subject_days[subject] = set()
            subject_days[subject].add(day_index)
    
    # Calculate the total number of days for all subjects
    total_days = sum(len(days) for days in subject_days.values())
    return total_days

def order_schedules_by_subject_day(schedules):
    """
    Orders the schedules by prioritizing those that have classes of the same subject on the same day.
    """
    # Sort schedules based on the total number of days a subject appears
    sorted_schedules = sorted(schedules, key=count_subject_days)
    return sorted_schedules

def main():
    schedules = []
    original_schedules = []
    is_first_time = True
    user_choice = ""
    global filter_settings

    while user_choice != "0":
        print("\n\n ||| SUPER TTS |||\n")
        print("Main Menu\n")
        print(f"Number of schedules: {len(schedules)}")
        print("\nChoose an option:")
        user_choice = input("1 - Calculate all schedules\n2 - Set filters\n3 - Output to Excel\n4 - Order schedules by subject days\n0 - Quit\n\nYour choice: ")

        if user_choice == "1":
            if is_first_time:
                print("\nCalculating all schedules...\n")
                schedules = generate_schedules(aulas, ["AED", "FII", "BD", "LDTS", "SO"], [[], [], [], [], []])
                original_schedules = copy.deepcopy(schedules)
                is_first_time = False
                print("Schedules generated successfully!\n")
            else:
                schedules = copy.deepcopy(original_schedules)
                filter_settings = {
                    "free_days": False,
                    "end_time": DEFAULT_END_TIME,
                    "professor": ""
                }
                print("\nSchedules and filter settings reset successfully!\n")

        elif user_choice == "2":
            filter_choice = ""
            while filter_choice != "0":
                print("\n\nFilter Options\n")
                print(f"Current settings:")
                print(f"- Free days: {'Yes' if filter_settings['free_days'] else 'No'}")
                print(f"- End time: {filter_settings['end_time']}")
                print(f"- Professor: {filter_settings['professor'] if filter_settings['professor'] else 'Not set'}")
                print(f"\nNumber of schedules: {len(schedules)}")
                print("\nChoose an option:")
                filter_choice = input('1 - Toggle free days\n2 - Set end time\n3 - Set professor\n4 - Apply filters\n9 - Reset all\n0 - Back to main menu\n\nYour choice: ')

                if filter_choice == "1":
                    filter_settings["free_days"] = not filter_settings["free_days"]
                    print(f"\nFree days filter {'enabled' if filter_settings['free_days'] else 'disabled'}")
                
                elif filter_choice == "2":
                    new_end_time = input(f'\nCurrent end time: {filter_settings["end_time"]}\nEnter new end time: ')
                    filter_settings["end_time"] = new_end_time
                    print(f"\nEnd time set to: {new_end_time}")

                elif filter_choice == "3":
                    new_professor = input("\nEnter the professor's name: ")
                    filter_settings["professor"] = new_professor
                    print(f"\nProfessor filter set to: {new_professor}")

                elif filter_choice == "4":
                    schedules = apply_filters(original_schedules, filter_settings)
                    print(f"\nFilters applied. Number of schedules after filtering: {len(schedules)}")

                elif filter_choice == "9":
                    filter_settings = {
                        "free_days": False,
                        "end_time": DEFAULT_END_TIME,
                        "professor": ""
                    }
                    schedules = copy.deepcopy(original_schedules)
                    print("\nFilter settings and schedules reset successfully.")

                elif filter_choice == "0":
                    break

        elif user_choice == "3":
            output_schedules_to_excel(schedules)
            print("\nSchedules exported to Excel successfully!\n")
        
        elif user_choice == "4":
            schedules = order_schedules_by_subject_day(schedules)
            print("\nSchedules ordered by subject days successfully!\n")

        elif user_choice == "0":
            print("\nThank you for using SUPER TTS. Goodbye!")
            break

if __name__ == "__main__":
    main()